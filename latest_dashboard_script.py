'''
Authors: Kharesa-Kesa and Jose
This script will attempt to replicate the actions of the spreadsheet generating the all_stations tab
'''

# from asyncio.unix_events import _UnixSelectorEventLoop
# from email import header
#from macpath import split
from operator import index
from matplotlib.pyplot import polar
import pandas as pd, numpy as np, glob, ast, openpyxl as xl, shutil, pyodbc, random, datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


def get_orr_step_free_category(search_value):
    orr = {'B': 'Bottom', 'B2': 'Bottom', 'B3': 'Middle', 'C': 'Top'}

    return orr[search_value]


def get_connectivity_journeys_matrix(search_value):
    e = {'TopTop': 1, 'TopMiddle': 2, 'TopBottom': 3, 'MiddleTop': 1, 'MiddleMiddle': 2, 'MiddleBottom': 3,
         'BottomTop': 2, 'BottomMiddle': 3, 'BottomBottom': 5}

    return e[search_value]


def get_updated_stations():
    input_path = r"C:\Users\Kharesa-Kesa.Spencer\OneDrive - Arup\Projects\Network Rail Accessibility case\matrices\Input template.csv"    
    input_df = pd.read_csv(input_path)

    scenario_tag = str(input_df.Scenario_Name.iat[0])
    input_df.drop(columns='Scenario_Name')

    return input_df, scenario_tag


def get_mobility_isolation_matrix(search_value):
    e = {'TopTop': 1, 'TopMiddle': 1, 'TopBottom': 2, 'MiddleTop': 1, 'MiddleMiddle': 2, 'MiddleBottom': 3,
         'BottomTop': 2, 'BottomMiddle': 3, 'BottomBottom': 3}

    return e[search_value]

def get_list_col():
    list_cols = ['Station_Name', 'Unique_Code', 'Station_Facility_Owner',
                 'Network_Rail_Region', 'ORR_Step_Free_Category', 'DfT_Category',
                 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)',
                 '2018/2019_ORR_Total_Entries/Exits_and_Interchange',
                 '2019_Journeys_to_an_accessible_destination',
                 '2019_Journeys_from_an_accessible_origin',
                 '2019_Total_Unlocked_Journeys', '2019_Potential_Unlocked_Rank',
                 '2019_Unlocked_Journeys_Percentile',
                 '2019_Unlocked_Journeys_Matrix_Outcome',
                 '2019_Connectivity_(count_of_stations_directly_served)',
                 '2019_Connectivity_Rank', '2019_Connectivity_Percentile',
                 '2019_Connectivity_Matrix_Outcome',
                 'Connectivity_and_Journeys_Matrix_Outcome',
                 'Connectivity_and_Journeys_Matrix_Outcome.1', 'Mobility_Score',
                 'Isolation_(1_if_no_Cat_A_in_20_mins_drive_isochrone)',
                 'Additional_Flags', 'Original_Isolation_Score',
                 'Revisited_Isolation_score', 'Mobility/Isolation',
                 'Isolation_and_Current_Access_Matrix_Outcome', 'Socioeconomic_Flags',
                 'Socioeconomic_classification', 'Local_Impact_Score',
                 'Local_Impact_Classification', 'Socioecon_/_Local_Impact',
                 'Socioeconomic_/_Local_Matrix_outcome', 'Average_of_two_scores',
                 'Score__without_modifier', 'Base_score_+_modifiers',
                 'Score_with_modifier', 'Footfall_Modifier', 'Final_Outcome', 'Change',
                 'Region_and_Final_Score', 'Journeys_and_Final_Score',
                 'Region_and_Local_Factor', 'Score_Change']

    one = ['Station_Name', 'Unique_Code', 'Station_Facility_Owner',
           'Network_Rail_Region', 'ORR_Step_Free_Category', 'DfT_Category',
           'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)',
           '2018/2019_ORR_Total_Entries/Exits_and_Interchange',
           '2019_Journeys_to_an_accessible_destination',
           '2019_Journeys_from_an_accessible_origin',
           '2019_Total_Unlocked_Journeys', '2019_Potential_Unlocked_Rank',
           '2019_Unlocked_Journeys_Percentile',
           '2019_Unlocked_Journeys_Matrix_Outcome',
           '2019_Connectivity_(count_of_stations_directly_served)',
           '2019_Connectivity_Rank', '2019_Connectivity_Percentile']

    two = ['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)', '2019_Connectivity_Matrix_Outcome',
           'Connectivity_and_Journeys_Matrix_Outcome',
           'Connectivity_and_Journeys_Matrix_Outcome.1', 'Mobility_Score',
           'Isolation_(1_if_no_Cat_A_in_20_mins_drive_isochrone)',
           'Additional_Flags', 'Original_Isolation_Score',
           'Revisited_Isolation_score', 'Mobility/Isolation',
           'Isolation_and_Current_Access_Matrix_Outcome', 'Socioeconomic_Flags',
           'Socioeconomic_classification', 'Local_Impact_Score',
           'Local_Impact_Classification', 'Socioecon_/_Local_Impact',
           'Socioeconomic_/_Local_Matrix_outcome', 'Average_of_two_scores',
           'Score__without_modifier', 'Base_score_+_modifiers',
           'Score_with_modifier', 'Footfall_Modifier', 'Final_Outcome', 'Change',
           'Region_and_Final_Score', 'Journeys_and_Final_Score',
           'Region_and_Local_Factor', 'Score_Change']

    three = ['Isolation_(1_if_no_Cat_A_in_20_mins_drive_isochrone)',
             'Additional_Flags', 'Original_Isolation_Score',
             'Revisited_Isolation_score', 'Mobility/Isolation',
             'Isolation_and_Current_Access_Matrix_Outcome', 'Socioeconomic_Flags',
             'Socioeconomic_classification', 'Local_Impact_Score',
             'Local_Impact_Classification', 'Socioecon_/_Local_Impact',
             'Socioeconomic_/_Local_Matrix_outcome', 'Average_of_two_scores',
             'Score__without_modifier', 'Base_score_+_modifiers',
             'Score_with_modifier', 'Footfall_Modifier', 'Final_Outcome', 'Change',
             'Region_and_Final_Score', 'Journeys_and_Final_Score',
             'Region_and_Local_Factor', 'Score_Change']

    return three

def get_DfT_Num_Cat(search_value):

    if search_value.isna():
        return None

    else:

        e = {'A':6, 'B':5, 'C':4, 'D':3, 'E':2, 'F':1, 'NaN': None}
        return e[search_value]

def input_OD_Matrix():
    # inputs
    # connect to the access database
    conn = pyodbc.connect(
        r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\Kharesa-Kesa.Spencer\OneDrive - Arup\Projects\Network Rail Accessibility case\matrices\MOIRAOD.accdb;')

    query = 'select * from ODMatrixAfA'
    OD_df = pd.read_sql(query, conn)

    return OD_df


def map_input_stations(OD_df, base_df, input_df):
    # Add numerical score based on a dictionary
    my_dict = {'0': 0, 'A': 1, 'B': 2, 'B1': 3, 'B2': 4, 'B3': 5, 'C': 6, 'Null': -1}
    origin_score = OD_df.AfAOrigin.map(my_dict)
    destination_score = OD_df.AfADest.map(my_dict)
    OD_df['origin_score'] = origin_score
    OD_df['destination_score'] = destination_score

    # Remove OD pairs where the score is 0 or Null. Save the number of removed records for traceability
    dropped_origins = len(OD_df[(OD_df.origin_score < 1) | (OD_df.origin_score.isna())])
    dropped_destinations = len(OD_df[(OD_df.destination_score < 1) | (OD_df.destination_score.isna())])
    OD_df = OD_df.drop(OD_df[(OD_df.origin_score < 1) | (OD_df.origin_score.isna())].index)
    OD_df = OD_df.drop(OD_df[(OD_df.destination_score < 1) | (OD_df.destination_score.isna())].index)

    print(f"A total of {dropped_origins:n} origins and {dropped_destinations:n} destinations were invalid and not included in the analysis.")

    # Categorize the journeys by looking at the maximum numerical score of the origin and destination stations
    jny_score = np.maximum(OD_df.origin_score, OD_df.destination_score)

    # Add the list as a new column to the existing dataframe
    OD_df['jny_score'] = jny_score
    my_dict_2 = {1: 'A', 2: 'B', 3: 'B1', 4: 'B2', 5: 'B3', 6: 'C'}
    jny_category = OD_df.jny_score.map(my_dict_2)
    OD_df['jny_category'] = jny_category
    OD_df['Total_Journeys'] = OD_df['STDJOURNEYS'] + OD_df['1stJOURNEYS']
    # print(base_df.tail(10))

    ## TEST
    OD_df.isna().sum()

    # base_df.info()
    #v1 = OD_df.groupby("jny_category").Total_Journeys.sum()
    # print(v1)
    # print(type(v1))
    # print(base_df.tail(10))

    # this method does all the calculating the new categories

    # Import stations to be upgraded
    upgrade_list = input_df

    # In anticipation of the new score the columm is set to None
    base_df['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = None

    for tlc in upgrade_list.TLC:
        if str(tlc) == 'nan':
            continue
        # Update category. It is necessary to use the .item() method to the Series ,
        new_category = upgrade_list.loc[upgrade_list.TLC == tlc, 'New_Category'].item()

        # Update origin category
        OD_df.loc[OD_df.OriginTLC == str(tlc), 'AfAOrigin'] = new_category
        # Update destination category
        OD_df.loc[OD_df.DestinationTLC == str(tlc), 'AfADest'] = new_category

        if tlc in base_df.values:
            # Update ORR_Step_Free_Category
            base_df.loc[base_df.Unique_Code == str(tlc), 'ORR_Step_Free_Category'] = new_category
            # setting their indicator to 0 as they are now accessible
            base_df.loc[base_df.Unique_Code == str(tlc), 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 0

    # Update origin score
    OD_df.origin_score = OD_df.AfAOrigin.map(my_dict)
    # Update destination score
    OD_df.destination_score = OD_df.AfADest.map(my_dict)
    # Update journey score
    OD_df.jny_score = np.maximum(OD_df.origin_score, OD_df.destination_score)
    # Update journey category
    OD_df.jny_category = OD_df.jny_score.map(my_dict_2)
    # Concat the 2 categories together
    OD_df['concat_categories'] = OD_df.AfAOrigin + OD_df.AfADest

    New_ODMatrix = OD_df
    New_ODMatrix.drop(axis=1,columns=['origin_score', 'destination_score', 'jny_score', 'jny_category'], inplace=True)

    # dataframes where only the origin or the destination are accessible
    OD_df_ass_origin = OD_df.loc[(OD_df.AfAOrigin == 'A') | (OD_df.AfAOrigin == 'B1')]
    OD_df_ass_destination = OD_df.loc[(OD_df.AfADest == 'A') | (OD_df.AfADest == 'B1')]

    grouped_origin_df = (OD_df_ass_origin.groupby(["OriginTLC", "AfAOrigin"])["Total_Journeys"].sum()).to_frame()
    grouped_origin_df.reset_index(inplace=True)

    grouped_destination_df = (
        OD_df_ass_destination.groupby(["DestinationTLC", "AfADest"])["Total_Journeys"].sum()).to_frame()
    grouped_destination_df.reset_index(inplace=True)


    # Placing the total journey grouped values into
    for code in base_df.Unique_Code:
        if str(tlc) == 'nan':
            continue

        if str(code) in grouped_origin_df.values:
            total_jo = grouped_origin_df.loc[grouped_origin_df.OriginTLC == str(code), 'Total_Journeys'].item()
            base_df.loc[base_df.Unique_Code == str(code), '2019_Journeys_from_an_accessible_origin'] = total_jo

        if str(code) in grouped_destination_df.values:
            total_jd = grouped_destination_df.loc[
                grouped_destination_df.DestinationTLC == str(code), 'Total_Journeys'].item()
            base_df.loc[base_df.Unique_Code == str(code), '2019_Journeys_to_an_accessible_destination'] = total_jd

    return base_df, grouped_origin_df, grouped_destination_df, New_ODMatrix


def get_new_categories_set_jrnys(base_df, input_df):
    # set the OD matrix, either from the access database or from the csv
    OD_df = input_OD_Matrix()

    base_df, grouped_origin_df, grouped_destination_df, New_ODMatrix = map_input_stations(OD_df, base_df, input_df)

    # so now the base df and the od df both have updated station categories from the input template
    # next is adding the column codes and the journey stats, mapping from the OD_df to the base df
    # if a station is accessible they're changed to 0 and if the column is empty its set to NaN
    # Did it manually due to error
    base_df.loc[base_df.ORR_Step_Free_Category == 'A', 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 0
    base_df.loc[base_df.ORR_Step_Free_Category == 'B1', 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 0
    base_df.loc[base_df.ORR_Step_Free_Category == 'B', 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 1
    base_df.loc[base_df.ORR_Step_Free_Category == 'B2', 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 1
    base_df.loc[base_df.ORR_Step_Free_Category == 'B3', 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 1
    base_df.loc[base_df.ORR_Step_Free_Category == 'C', 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = 1

    base_df.loc[base_df.ORR_Step_Free_Category.isna(), 'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] = np.NaN

    unlock_jrny = []
    unlocked_jrny_perc = []
    connc = []
    connc_rank = []
    connc_rank_perc = []
    connc_jrnys_matrx = []

    matrix_outcome_list = []
    matrix_outcome_list_2 = []

    for index, row in base_df.iterrows():

        if row['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 1:

            unlock_jrny.append(
                int(row['2019_Journeys_to_an_accessible_destination'] + row['2019_Journeys_from_an_accessible_origin']))
            connc.append(int(row['2019_Connectivity_(count_of_stations_directly_served)']))

            # setting this by looping through df with if statements
        else:
            unlock_jrny.append(None)
            connc.append(None)

    base_df['2019_Total_Unlocked_Journeys'] = unlock_jrny
    unlock_jrny_s = pd.Series(unlock_jrny)
    base_df['2019_Potential_Unlocked_Rank'] = unlock_jrny_s.rank(ascending=False)
    base_df['2019_Unlocked_Journeys_Percentile'] = unlock_jrny_s.rank(ascending=False, pct=True)
    base_df['2019_Connectivity_(count_of_stations_directly_served)'] = connc
    connc_s = pd.Series(connc)
    base_df['2019_Connectivity_Rank'] = connc_s.rank(ascending=False)
    base_df['2019_Connectivity_Percentile'] = connc_s.rank(ascending=False, pct=True)

    for index, row in base_df.iterrows():
        val = row['2019_Unlocked_Journeys_Percentile']
        val_2 = row['2019_Connectivity_Percentile']

        if row['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 1:

            if val < 0.33:
                matrix_outcome_list.append('Top')

            elif val < 0.66:
                matrix_outcome_list.append('Middle')

            else:
                matrix_outcome_list.append('Bottom')

        else:
            matrix_outcome_list.append('')

        if row['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 1:

            if val_2 < 0.33:
                matrix_outcome_list_2.append('Top')

            elif val_2 < 0.66:
                matrix_outcome_list_2.append('Middle')

            else:
                matrix_outcome_list_2.append('Bottom')

        else:
            matrix_outcome_list_2.append('')

    base_df['2019_Unlocked_Journeys_Matrix_Outcome'] = matrix_outcome_list
    base_df['2019_Connectivity_Matrix_Outcome'] = matrix_outcome_list_2
    base_df['Connectivity_and_Journeys_Matrix_Outcome'] = base_df['2019_Unlocked_Journeys_Matrix_Outcome'] + base_df[
        '2019_Connectivity_Matrix_Outcome']

    # Matrix
    for index, row in base_df.iterrows():
        mob = str(row['Connectivity_and_Journeys_Matrix_Outcome'])

        if mob == 'nan' or mob == '':
            continue
        else:
            base_df.loc[base_df[
                            'Connectivity_and_Journeys_Matrix_Outcome'] == mob, 'Connectivity_and_Journeys_Matrix_Outcome.1'] = get_connectivity_journeys_matrix(
                mob)
          

    return base_df, grouped_origin_df, grouped_destination_df, New_ODMatrix


def set_mobility_isolation_score(updated_cats_and_jrnys, alt_any_df, input_df):
    # COLUMN V

    # Matrix
    for index, row in updated_cats_and_jrnys.iterrows():

        if row['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 1:
            osfc = str(row['ORR_Step_Free_Category'])

            updated_cats_and_jrnys.loc[updated_cats_and_jrnys['Unique_Code'] == str(
                row['Unique_Code']), 'Mobility_Score'] = get_orr_step_free_category(osfc)

        else:
            updated_cats_and_jrnys.loc[
                updated_cats_and_jrnys['Unique_Code'] == str(row['Unique_Code']), 'Mobility_Score'] = None

    # getting the input stations
    list_of_changed_stns = list(input_df['TLC'])

    # COLUMN Z
    # creating a blank dataframe with the same columns as the alt_any sheet table
    stns_df = pd.DataFrame(columns=['Station_Code', 'Station_Name', 'Region', 'ID2', 'Station__2'])

    # looping through the list of inputted stations
    # for every station thats in the in the dataframe
    # check if this is the first time the dataframe is being appended (through length) if not then
    # append all the rows where the station code equals the one being looped through
    for station in list_of_changed_stns:

        if station in alt_any_df.values:

            if len(stns_df) == 0:
                stns_df = alt_any_df[alt_any_df.Station_Code == station]

            else:
                temp_df = alt_any_df[alt_any_df.Station_Code == station]
                stns_df = pd.concat([stns_df, temp_df], axis=0, )

    # a list of all the station codes of the stations within 20 of all the newly upgraded stations (using set to remove duplicates)
    target_stns = list(set(stns_df.ID2))

    # now we pivot over top the all station sheet to assign bottom to all these stations

    for target in target_stns:

        if target in updated_cats_and_jrnys.values:
            updated_cats_and_jrnys.loc[(updated_cats_and_jrnys.Unique_Code == target) & (updated_cats_and_jrnys[
                                                                                             'Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 1), 'Revisited_Isolation_score'] = 'Bottom'

    # COLUMN Y

    updated_cats_and_jrnys['Mobility/Isolation'] = updated_cats_and_jrnys['Original_Isolation_Score'] + \
                                                   updated_cats_and_jrnys['Revisited_Isolation_score']

    # Matrix
    for index, row in updated_cats_and_jrnys.iterrows():
        mob = str(row['Mobility/Isolation'])

        if mob == 'nan':
            continue
        else:
            updated_cats_and_jrnys.loc[updated_cats_and_jrnys[
                                           'Mobility/Isolation'] == mob, 'Isolation_and_Current_Access_Matrix_Outcome'] = get_mobility_isolation_matrix(
                mob)

    # dropping final duplicates before returning
    updated_cats_and_jrnys = updated_cats_and_jrnys.loc[:, ~updated_cats_and_jrnys.columns.duplicated()].copy()

    return updated_cats_and_jrnys


def blanking_rows(updated_mobility_and_isolation):
    list_of_cols = get_list_col()

    for index, row in updated_mobility_and_isolation.iterrows():

        for col in list_of_cols:

            if row['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 0:
                updated_mobility_and_isolation.loc[updated_mobility_and_isolation['Inaccessible_(1_if_not_Step_Free_Cat._A_or_B1)'] == 0, col]=None

    updated_mobility_and_isolation['Dft_Category'] = updated_mobility_and_isolation['Dft_Category'].fillna('')

    return updated_mobility_and_isolation


def into_stepfree_spreadsheet(final_df, grouped_origin_df, grouped_destination_df, path_of_spreadsh, scenario_tag):

    #Final_df is the all_stations sheet here with the new updated station cateogries in in
    #grouped origin and destination are grouped dfs of the total journeys grouped by station
    #this method is to write to the new spreadsheet clones


    #clones spreadsheet as to not affect the original when writing to the sheet
    original = path_of_spreadsh
    clone = r"C:\Users\Kharesa-Kesa.Spencer\OneDrive - Arup\Projects\Network Rail Accessibility case\CSV WORK\Step Free Scoring_JDL_v3.00_clone.xlsx"
    target = r"C:\Users\Kharesa-Kesa.Spencer\OneDrive - Arup\Projects\Network Rail Accessibility case\CSV WORK\Step Free Scoring_JDL_v3.00"+scenario_tag+'.xlsx'

    #copying file files
    shutil.copyfile(original, clone)

    #Slimming down the clone
    workbook=xl.load_workbook(clone)
    sheet_names=workbook.sheetnames
    sheet_names.remove('Dashboard')


    for sheet in sheet_names:
        del workbook[sheet]

    workbook.save(target)

    #reading from the spreadsheet
    st_cat_df = pd.read_excel(clone, sheet_name = "St_Cat", engine='openpyxl')
    st_cat_df.rename(columns={'CRS Code': 'CRS_Code', 'Station Name (MOIRA Name)': 'Station_Name'},inplace=True)

    st_cat_df = st_cat_df.loc[:,~st_cat_df.columns.duplicated()].copy()


    # Next steps upgrading the data in the clone to this current scenario:
    # 1. find the stations to be upgraded.
    for code in final_df.Unique_Code:
        if str(code) == 'nan':
            continue

        #if the current code is in the station category spreadsheet then upgrade st_cat_df with new category
        if code in st_cat_df.values:
            new_category = final_df.loc[final_df.Unique_Code == code, 'ORR_Step_Free_Category'].item()
            st_cat_df.loc[st_cat_df.CRS_Code == str(code), 'Including CP6 AfA'] = new_category

    
    #setting all Cat A as None

    grouped_origin_df.loc[grouped_origin_df.AfAOrigin=='A', 'Total_Journeys'] = None
    grouped_origin_df.loc[grouped_origin_df.AfAOrigin=='B1', 'Total_Journeys'] = None 

    grouped_destination_df.loc[grouped_destination_df.AfADest=='A', 'Total_Journeys'] = None 
    grouped_destination_df.loc[grouped_destination_df.AfADest=='B1', 'Total_Journeys'] = None 
    
    #export back to csv
    with pd.ExcelWriter(target, mode="a",engine="openpyxl") as writer:
        
        final_df.to_excel(writer, sheet_name="All Stations", index=False)
        st_cat_df.to_excel(writer, sheet_name="St_Cat", index=False)
        grouped_origin_df.to_excel(writer, sheet_name="Inaccessible O Accessi D")
        grouped_destination_df.to_excel(writer, sheet_name="Accessible O Inaccessi D")
        
    #done

def output_to_log(input_df, scenario_tag ):
    
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    
    with open(r"C:\Users\Kharesa-Kesa.Spencer\OneDrive - Arup\Projects\Network Rail Accessibility case\CSV WORK\scenarios_output_log.txt", "a+") as file_object:
        # Move read cursor to the start of file.
        file_object.seek(0)
        # If file is not empty then append '\n'
        data = file_object.read(100)
        if len(data) > 0 :
            file_object.write("\n")
        # Append text at the end of file
        file_object.write("\n")
        line = 'Scenario number: ' + scenario_tag + ' run at ' + dt_string
        file_object.write(line)
        file_object.write("\n")
        file_object.write(str(input_df))
        file_object.write("\n")

 def make_kepler_input(final_df, path_of_spreadsh, scenario_tag):

    #naming this dataframe kepler as it will serve as the stations.csv file for kepler
    kepler = pd.read_excel(path_of_spreadsh, sheet_name="Coordinates", header=2, usecols="B:F", engine='openpyxl')

    #running over the final_df dafeframe and adding the new categories and other columns to it 
    for code in final_df.Unique_Code:
        if str(code) == 'nan':
            continue
        if code in kepler.values:

            # Update category. It is necessary to use the .item() method to the Series ,
            kepler.loc[kepler.CRSCode == str(code), 'Step Free Category'] = final_df.loc[final_df.Unique_Code == code, 'ORR_Step_Free_Category'].item()
            #kepler.loc[kepler.CRSCode == str(code), 'DfT Category']  = final_df.loc[final_df.Unique_Code == code, 'Dft_Category'].item()
            kepler.loc[kepler.CRSCode == str(code), 'Journeys Unlocked']  = final_df.loc[final_df.Unique_Code == code, '2019_Total_Unlocked_Journeys'].item()
            kepler.loc[kepler.CRSCode == str(code), 'Connectivity Rank']  = final_df.loc[final_df.Unique_Code == code, '2019_Connectivity_Rank'].item()
            kepler.loc[kepler.CRSCode == str(code), 'Isolation Score']  = final_df.loc[final_df.Unique_Code == code, 'Isolation_and_Current_Access_Matrix_Outcome'].item()
            kepler.loc[kepler.CRSCode == str(code), 'Final Matrix Outcome']  = final_df.loc[final_df.Unique_Code == code, 'Final_Outcome'].item()
            #kepler.loc[kepler.CRSCode == str(code), 'DfT Category A to F']  = get_DfT_Num_Cat(final_df.loc[final_df.Unique_Code == code, 'Dft_Category'].item())

    #Exporting the file as a csv
    csv_outpath = 'Stations_sc_'+scenario_tag
    kepler.to_csv(csv_outpath)



#Pseudo-Main

path_of_spreadsh =  r"C:\Users\Kharesa-Kesa.Spencer\OneDrive - Arup\Projects\Network Rail Accessibility case\CSV WORK\Step Free Scoring_JDL_v3.00.xlsx"
base_df = pd.read_excel(path_of_spreadsh, sheet_name="All Stations", header=2, usecols="B:AS", engine='openpyxl')
alt_any = pd.read_excel(path_of_spreadsh, sheet_name="Alt_Any_20", header=4, usecols="B:F", engine='openpyxl')
input_df, scenario_tag = get_updated_stations()


base_df.columns = [c.replace(' ', '_') for c in base_df.columns]
alt_any.columns = [c.replace(' ', '_') for c in alt_any.columns]

updated_cats_and_jrnys, grouped_origin_df, grouped_destination_df, New_ODMatrix = get_new_categories_set_jrnys(base_df, input_df)
#
updated_mobility_and_isolation = set_mobility_isolation_score(updated_cats_and_jrnys, alt_any, input_df)
#
final_df = blanking_rows(updated_mobility_and_isolation)

output_to_log(input_df, scenario_tag)

into_stepfree_spreadsheet(final_df, grouped_origin_df, grouped_destination_df, path_of_spreadsh, scenario_tag)

make_kepler_input(final_df, path_of_spreadsh, scenario_tag)
